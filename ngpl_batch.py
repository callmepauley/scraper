import lxml.html
import requests
import random
import time
import csv
import sys

from datetime import datetime, timedelta
from openpyxl import load_workbook
from io import BytesIO


def load_excel(file, locations, date):
    wb = load_workbook(filename=BytesIO(file))
    ws = wb.active

    output = []

    for row in ws.iter_rows():
        try:
            loc = int(row[0].value)
            if loc in locations:
                _r = [str(cell.value).strip() for cell in row]
                _r.insert(0, date)
                output.append(_r)
                if len(output) == len(locations):
                    break
        except:
            pass

    return output


def load_page(url, date):
    session = requests.session()
    main_page = lxml.html.document_fromstring(session.get(url).text)

    # random data

    __VIEWSTATE = main_page.xpath('//input[@id="__VIEWSTATE"]/@value')[0]
    __VIEWSTATEGENERATOR = main_page.xpath('//input[@id="__VIEWSTATEGENERATOR"]/@value')[0]
    __EVENTVALIDATION = main_page.xpath('//input[@id="__EVENTVALIDATION"]/@value')[0]
    __ASYNCPOST = 'true'
    _clientstate = '|0|01{0}-0-0-0-0||[[[[]],[],[]],[{{}},[]],"01{0}-0-0-0-0"]'.format(date)  # - Y, M, D
    _location = 'rbDelivery'
    _cycle = '[[[[null,null,null,null,null,null,null,0,null,null,null,null,null,null,null,null,null,null,null,null,null,null,null,"BEST%20AVAILABLE",null,null,null,null,null,null,null,null,null,null,null,null,null,0,null,null,null,-1,null,null,null,null,null,null,null,null]],[],null],[{},[{}]],null]'
    # _cycle = '[[[[null,null,null,null,null,null,null,0,null,null,null,null,null,null,null,null,null,null,null,null,null,null,null,"BEST%20AVAILABLE",null,null,null,null,null,null,null,null,null,null,null,null,null,0,null,null,null,-1,null,null,null,null,null,null,null,null]],[],null],[{},[{}]],null]'
    _x = random.randrange(1, 80)
    _y = random.randrange(1, 30)

    data = {
        '__EVENTTARGET': '',
        '__EVENTARGUMENT': '',
        '__VIEWSTATE': __VIEWSTATE,
        '__VIEWSTATEGENERATOR': __VIEWSTATEGENERATOR,
        '__EVENTVALIDATION': __EVENTVALIDATION,
        '__ASYNCPOST': __ASYNCPOST,
        'ctl00_WebSplitter1_tmpl1_ContentPlaceHolder1_dtePickerBegin_clientState': _clientstate,
        'ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$location': _location,
        'ctl00_WebSplitter1_tmpl1_ContentPlaceHolder1_ddlCycleDD_clientState': _cycle,
        'ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.x': str(_x),
        'ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload.y': str(_y),
        'ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$DownloadDDL': 'EXCEL'
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:49.0) Gecko/20100101 Firefox/49.0",
    }

    xhr = session.post(url, data=data, headers=headers)

    if xhr.text.lstrip().startswith('<!DOCTYPE html'):
        retry = input('Something went wrong, retry? (Y/N): ')
        if retry.upper() == 'Y':
            load_page(url, date)
        else:
            return None
    else:
        return xhr.content


def save_output(data):
    with open('output.csv', 'w', newline='') as f:
        c = csv.writer(f, dialect='excel', quoting=csv.QUOTE_ALL)
        c.writerows(data)


def main():

    if len(sys.argv) == 3:
        try:
            min_d = max(0, min(89, int(sys.argv[1]), int(sys.argv[2])))
            max_d = min(90, max(1, int(sys.argv[1]), int(sys.argv[2])))
        except:
            min_d = 0
            max_d = 1
    elif len(sys.argv) == 2:
        min_d = 0
        try:
            max_d = min(90, max(1, int(sys.argv[1])))
        except:
            max_d = 1
    else:
        min_d = 0
        max_d = 1

    locations = [46622]
    # locations = [43978, 46622]

    output = []

    for d in range(min_d, max_d, 1):
        day = datetime.today() - timedelta(days=d)
        date = '-'.join([str(day.year), str(day.month), str(day.day)])
        bytefile = load_page('http://pipeline2.kindermorgan.com/Capacity/OpAvailPoint.aspx?code=NGPL', date)
        if bytefile is not None:
            day_data = load_excel(bytefile, locations, date)
            output.extend(day_data)
            delta = (time.perf_counter() / (d - min_d + 1))
            eta = int(((max_d - d) * delta) / 60)

            print("Retrieved {0} / {1} | {5} | Time elapsed: {2} s ({3} min) | ETA: {4} min".format(d + 1, max_d, int(time.perf_counter()), int(time.perf_counter() / 60), eta, date))
            time.sleep(60)
        else:
            print('Error downloading file.')

    save_output(output)
    print("Saved")

if __name__ == '__main__':
    main()
